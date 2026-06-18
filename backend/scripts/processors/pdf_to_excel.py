import os
from pathlib import Path
from .utils import openpyxl, extract_pdf_text

def pdf_to_excel(input_paths, output_path):
    if not input_paths:
        return
    pdf_text = extract_pdf_text(input_paths[0])
    data_mode = os.environ.get("PDF_TO_EXCEL_DATA", "all-tables").upper()
    sep = "," if os.environ.get("PDF_TO_EXCEL_SEPARATOR", "period").lower() == "comma" else "."
    detect = os.environ.get("PDF_TO_EXCEL_DETECT_NUM", "true").lower() == "true"

    if openpyxl:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        sheet = wb.active
        sheet.title = "PDF Data"

        # Header row styles
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E2D3D")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        alt_fill_1 = PatternFill("solid", fgColor="F4F6F8")
        alt_fill_2 = PatternFill("solid", fgColor="FFFFFF")

        # Parse the PDF text into table-like rows
        lines = [ln for ln in pdf_text.split("\n") if ln.strip()]

        # Try to detect column separator (2+ spaces, tabs, or pipes)
        import re
        parsed_rows = []
        for line in lines:
            if "|" in line:
                cells = [c.strip() for c in line.split("|") if c.strip()]
            elif "\t" in line:
                cells = [c.strip() for c in line.split("\t")]
            else:
                # Split on 2+ spaces
                cells = re.split(r"  +", line.strip())
                cells = [c.strip() for c in cells if c.strip()]
            if cells:
                parsed_rows.append(cells)

        if not parsed_rows:
            parsed_rows = [["No tabular data detected in PDF"]]

        # Calculate max columns
        max_cols = max(len(r) for r in parsed_rows)

        # Write metadata header
        sheet.cell(row=1, column=1, value=f"Mode: {data_mode} | Sep: '{sep}' | Detect Numeric: {detect}")
        sheet.cell(row=1, column=1).font = Font(name="Calibri", italic=True, size=9, color="888888")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(max_cols, 4))

        # Column headers row (Row 2)
        for c_idx in range(max_cols):
            cell = sheet.cell(row=2, column=c_idx + 1, value=f"Column {c_idx + 1}")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for r_idx, row in enumerate(parsed_rows):
            excel_row = r_idx + 3  # offset by 2 header rows
            alt_fill = alt_fill_1 if r_idx % 2 == 0 else alt_fill_2
            for c_idx in range(max_cols):
                val_str = row[c_idx] if c_idx < len(row) else ""
                # Attempt numeric detection
                cell_val = val_str
                if detect and val_str:
                    clean = val_str.replace(",", "").replace(sep, ".")
                    try:
                        cell_val = int(clean) if "." not in clean else float(clean)
                    except ValueError:
                        pass

                cell = sheet.cell(row=excel_row, column=c_idx + 1, value=cell_val)
                cell.fill = alt_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(name="Calibri", size=10)

                # Format numbers
                if isinstance(cell_val, float):
                    cell.number_format = '#,##0.00'
                elif isinstance(cell_val, int):
                    cell.number_format = '#,##0'

        # Auto-fit column widths (estimate)
        for c_idx in range(1, max_cols + 1):
            max_len = 0
            col_letter = get_column_letter(c_idx)
            for row_cells in sheet.iter_rows(min_col=c_idx, max_col=c_idx):
                for cell in row_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

        # Freeze the header row
        sheet.freeze_panes = "A3"

        wb.save(output_path)
    else:
        # Plain CSV fallback
        lines_out = [f"Mode: {data_mode} | Sep: '{sep}' | Detect: {detect}"]
        for line in pdf_text.split("\n"):
            lines_out.append(",".join(line.split()))
        Path(output_path).write_text("\n".join(lines_out), encoding="utf-8")

    print(f"Converted PDF to Excel sheet: {output_path}")
